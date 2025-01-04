import openpyxl

wb = openpyxl.load_workbook("people.xlsx")

ws = wb["people"]

cell = ws["A1"]
print(cell.value)


for row in ws.iter_rows(min_row=1,max_col=3,values_only=True):
    print(row)